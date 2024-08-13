
import configparser
import os
import argparse
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from pycarlo.core import Client, Query, Session

BATCH = 100
OUTPUT_FILE = "data_products_monitoring_coverage.xlsx"


def get_dp_summary(mc_client: Client, dp_uuid: str = None) -> dict:

    if dp_uuid:
        dps = {dp: {} for dp in list(dp_uuid.split(" "))}
    else:
        query = Query()
        query.get_data_products().__fields__("uuid", "is_deleted")
        dps = {dp.uuid: {} if not dp.is_deleted else '' for dp in mc_client(query).get_data_products}

    for dp in dps:
        cursor = None
        while True:
            get_data_product_summary_v2_query = f"""
            query getDataProductSummaryV2($dataProductId: UUID!, $upstreamLevels: Int) {{
                getDataProductV2(
                    dataProductId: $dataProductId
                    upstreamLevels: $upstreamLevels
                ) {{
                    uuid
                    name
                    monitored
                    warehouseUuids
                    tableCount
                    monitoredTableCount
                    assets(first: {BATCH} {f', after: "{cursor}"' if cursor else ''}) {{
                        pageInfo {{
                            hasNextPage
                            endCursor
                        }}
                        edges {{
                            node {{
                                displayName
                                objectType
                                mcon
                                upstreamDependenciesCount
                                isDeleted
                                importanceScore
                            }}
                        }}
                    }}
                }}
            }}"""
            summary = json.loads(json.dumps(mc_client(get_data_product_summary_v2_query,
                                                   variables={"dataProductId": dp, "upstreamLevels": 30})
                                            .get_data_product_v2, default=lambda o: o.__dict__))
            if dps.get(summary.get("uuid")):
                dps[summary["uuid"]]["assets"]["edges"].append(summary["assets"]["edges"])
            else:
                dps[summary["uuid"]] = summary

            if summary.get("assets").get("page_info").get("has_next_page"):
                cursor = summary["assets"]["page_info"]["end_cursor"]
            else:
                del dps[summary["uuid"]]["assets"]["page_info"]
                break

    return dps


def get_custom_monitors(mc_client: Client) -> tuple:

    monitors = {}
    monitors_per_table = {}
    skip_records = 0
    while True:
        query = Query()
        query.get_monitors(limit=BATCH, offset=skip_records).__fields__("uuid", "entities", "description", "monitor_type",
                                                                        "monitor_status", "monitor_run_status",
                                                                        "resource_id", "name")
        response = mc_client(query).get_monitors
        if len(response) > 0:
            for monitor in response:
                monitors[monitor.uuid] = monitor
                if monitor.monitor_type == "COMPARISON":
                    query = Query()
                    cr_query = query.get_custom_rule(rule_id=monitor.uuid)
                    cr_query.__fields__("uuid", "rule_type", "is_paused")
                    cr_query.queries(first=BATCH).edges.node.__fields__("entities")
                    rule = mc_client(query).get_custom_rule
                    monitors[monitor.uuid].entities.extend(edge.node.entities[0] for edge in rule.queries.edges if edge.node.entities[0] not in monitors[monitor.uuid].entities)
                    print()

                if monitor.entities:
                    for entity in monitor.entities:
                        if not monitors_per_table.get(entity):
                            monitors_per_table[entity] = [{"uuid": monitor.uuid, "description": monitor.description}]
                        else:
                            monitors_per_table[entity].append({"uuid": monitor.uuid, "description": monitor.description})

        skip_records += BATCH
        if len(response) < BATCH:
            break

    return monitors, monitors_per_table


def get_dp_tables(mc_client: Client, data_products: dict):

    _, table_monitors = get_custom_monitors(mc_client)
    for dp in data_products:
        for edge in data_products[dp]["assets"]["edges"]:
            get_connected_mcon_lineage_query = f"""
                query getConnectedMconLineage($mcons: [String]!, $levels: Int = 20) {{
                    getConnectedMconLineage(mcons: $mcons, levels: $levels) {{
                        connectedMcons {{
                            mcon
                        }}
                    }}
                }}"""
            res = (client(get_connected_mcon_lineage_query, variables={"mcons": [edge["node"]["mcon"]], "levels": 20})
                   .get_connected_mcon_lineage)
            tables_mcons = [asset["mcon"] for asset in json.loads(json.dumps(res.connected_mcons, default=lambda o: o.__dict__))]
            cursor = None
            while True:
                get_tables_query = f"""
                    query getTables($after: String, $first: Int, $isDeleted: Boolean, $mcons: [String]) {{
                        getTables(after: $after, first: $first, isDeleted: $isDeleted, mcons: $mcons) {{
                            pageInfo {{
                                hasNextPage
                                endCursor
                            }}
                            edges {{
                                node {{
                                    mcon
                                    fullTableId
                                    tableType
                                    isMonitored
                                    isMuted
                                    importanceScore
                                }}
                            }}
                        }}
                    }}"""
                res = (client(get_tables_query, variables={"first": BATCH, "last": cursor, "mcons": tables_mcons,
                                                           "isDeleted": False}).get_tables)
                res_json = json.loads(json.dumps(res, default=lambda o: o.__dict__))
                custom_monitor_count = 0
                for asset in res_json["edges"]:
                    if table_monitors.get(asset["node"]["full_table_id"]):
                        asset["node"]["custom_monitors"] = table_monitors[asset["node"]["full_table_id"]]
                        custom_monitor_count += 1

                    if not edge["node"].get("tables"):
                        edge["node"]["tables"] = [asset["node"]]
                    else:
                        edge["node"]["tables"].append(asset["node"])

                edge["node"]["custom_monitor_count"] = custom_monitor_count

                if res_json["page_info"]["has_next_page"]:
                    cursor = res_json["page_info"]["end_cursor"]
                else:
                    break


def create_dataframe(data_dict):
    wb = Workbook()
    del wb['Sheet']
    rows = []
    for key, value in data_dict.items():
        uuid = value['uuid']
        name = value['name']
        table_count = value['table_count']
        monitored_table_count = value['monitored_table_count']
        monitored_percentage = f"{round((monitored_table_count / table_count) * 100, 2)} %" if table_count > 0 else "0 %"

        for edge in value['assets']['edges']:
            node = edge['node']
            display_name = node['display_name']
            upstream_dependencies_count = node['upstream_dependencies_count']
            custom_monitored_count = node['custom_monitor_count']
            custom_monitored_percentage = f"{round((custom_monitored_count / upstream_dependencies_count) * 100, 2)} %" if upstream_dependencies_count > 0 else "0 %"

            for table in node['tables']:
                full_table_id = table['full_table_id']
                table_type = table['table_type']
                table_importance_score = table['importance_score']
                table_mcon = table['mcon']

                # If there are no monitors, create a single row for the table
                if not table.get('custom_monitors'):
                    rows.append([
                        name, table_count, monitored_table_count, monitored_percentage,
                        display_name, upstream_dependencies_count, custom_monitored_count, custom_monitored_percentage,
                        full_table_id, table_mcon, table_type, False,
                        table_importance_score, None, None
                    ])
                else:
                    for monitor in table.get('custom_monitors'):
                        monitor_uuid = monitor['uuid']
                        description = monitor['description']

                        rows.append([
                            name, table_count, monitored_table_count, monitored_percentage,
                            display_name, upstream_dependencies_count, custom_monitored_count, custom_monitored_percentage,
                            full_table_id, table_mcon, table_type, True,
                            table_importance_score, monitor_uuid, description
                        ])

        # Creating the DataFrame
        columns = [
            'Name', 'Table Count', 'Monitored Table Count', 'Monitored %',
            'Table/Report', 'Upstream Dependencies Count', 'Custom Monitored Count', 'Custom Monitored %', 'Full Table ID',
            'Table MCON', 'Table Type', 'Custom Monitored',
            'Table Importance Score', 'Monitor UUID', 'Monitor Description'
        ]
        df = pd.DataFrame(rows, columns=columns)
        wb.create_sheet(name)
        ws = wb[name]
        ws.append(["Data Products Monitoring Coverage"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1)
        title_cell.fill = PatternFill(start_color='0c5395', end_color='0c5395', fill_type='solid')
        title_cell.font = Font(name='Roboto', size=18, color='ffffff', bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Append the DataFrame to the worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Adjust column widths to fit the text
        for col in ws.columns:
            max_length = 0
            col_letter = col[1].column_letter  # Get the column letter
            for i, cell in enumerate(col[1:ws.max_row + 1]):
                if i == 0:
                    cell.font = Font(name='Roboto', size=16, bold=True, color='0c5395')
                    cell.fill = PatternFill(start_color='f3f3f3', end_color='f3f3f3', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = Font(name='Roboto', size=14)

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width

        # Merge cells for each column where applicable
        for col in [1, 2, 3, 4, 5, 9, 10]:
            merge_cells(ws, 3, ws.max_row, col)
        for col in [6, 7, 8, 11, 12, 13]:
            merge_upstream_dependencies(ws, 3, ws.max_row, col)

        # Add hyperlinks
        full_table_id_col_idx = columns.index('Full Table ID') + 1
        table_mcon_col_idx = columns.index('Table MCON') + 1
        monitor_description_col_idx = columns.index('Monitor Description') + 1
        monitor_uuid_col_idx = columns.index('Monitor UUID') + 1
        for row in range(3, ws.max_row + 1):
            dp = ws.cell(row=row, column=1)
            if dp.value:
                dp.hyperlink = f"https://getmontecarlo.com/data-products/{uuid}"
                dp.style = "Hyperlink"
                dp.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
                dp.font = Font(size=50)

            table_mcon = ws.cell(row=row, column=table_mcon_col_idx)
            if table_mcon.value:
                table_id_cell = ws.cell(row=row, column=full_table_id_col_idx)
                table_id_cell.alignment = Alignment(vertical='center')
                if table_id_cell.value:
                    table_id_cell.hyperlink = f"https://getmontecarlo.com/assets/{table_mcon.value}"
                    table_id_cell.style = "Hyperlink"

            monitor_uuid = ws.cell(row=row, column=monitor_uuid_col_idx)
            if monitor_uuid.value:
                monitor_description_cell = ws.cell(row=row, column=monitor_description_col_idx)
                monitor_description_cell.hyperlink = f"https://getmontecarlo.com/monitors/{monitor_uuid.value}"
                monitor_description_cell.style = "Hyperlink"

            # Set columns to rotate text up
            for i in range(1, 9):
                cell = ws.cell(row=row, column=i)
                if cell.value:
                    cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
                    cell.font = Font(size=40, bold=True)

        # Hide Certain Columns
        hide_cols = ['Table MCON', 'Monitor UUID']
        hide_indices = [chr(ord('@')+columns.index(col) + 1) for col in hide_cols]
        for col in hide_indices:
            ws.column_dimensions[col].hidden = True

    wb.save(OUTPUT_FILE)


def merge_cells(ws, start_row, end_row, col):
    cell_value = ws.cell(row=start_row, column=col).value
    merge_start = start_row

    for row in range(start_row + 1, end_row + 1):
        next_cell_value = ws.cell(row=row, column=col).value

        if cell_value == next_cell_value:
            ws.cell(row=row, column=col, value=None)  # Clear the cell value
        else:
            if merge_start < row - 1:
                ws.merge_cells(start_row=merge_start, start_column=col, end_row=row - 1, end_column=col)
                ws.cell(row=merge_start, column=col).alignment = Alignment(vertical='center')

            cell_value = next_cell_value
            merge_start = row

    if merge_start < end_row:
        ws.merge_cells(start_row=merge_start, start_column=col, end_row=end_row, end_column=col)
        ws.cell(row=merge_start, column=col).alignment = Alignment(vertical='center')


def merge_upstream_dependencies(ws, start_row, end_row, col):
    merge_start = start_row

    while merge_start <= end_row:
        cell_value = ws.cell(row=merge_start, column=col).value
        if cell_value is not None:
            try:
                merge_end = merge_start
                while (merge_end + 1 <= end_row and
                       ws.cell(row=merge_end + 1, column=col).value == cell_value and
                       ws.cell(row=merge_end + 1, column=col - 1).value is None):
                    merge_end += 1
                if merge_start < merge_end:
                    ws.merge_cells(start_row=merge_start, start_column=col, end_row=merge_end, end_column=col)
                    ws.cell(row=merge_start, column=col).alignment = Alignment(vertical='center')

                merge_start = merge_end + 1
            except ValueError:
                merge_start += 1
        else:
            merge_start += 1


if __name__ == '__main__':

    # Capture Command Line Arguments
    parser = argparse.ArgumentParser(description='Track DP Unmonitored Tables')
    subparsers = parser.add_subparsers(dest='commands', required=True, metavar="<command>")
    parser._optionals.title = "Options"
    parser._positionals.title = "Commands"

    args = parser.parse_args()

    # Initialize variables
    profile = args.profile

    # Read token variables from CLI default's config path ~/.mcd/profiles.ini
    configs = configparser.ConfigParser()
    profile_path = os.path.expanduser("~/.mcd/profiles.ini")
    configs.read(profile_path)
    mcd_id_current = configs[profile]['mcd_id']
    mcd_token_current = configs[profile]['mcd_token']
    client = Client(session=Session(mcd_id=mcd_id_current, mcd_token=mcd_token_current))

    # Generate DP report
    dp_summary = get_dp_summary(client, args.data_product)
    get_dp_tables(client, dp_summary)
    create_dataframe(dp_summary)
    print("Done")
